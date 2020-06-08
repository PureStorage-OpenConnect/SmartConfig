#!/usr/bin/env python
# Project_Name    :FlashStack SmartConfig
# title           :reportmanager.py
# description     :Used for generating SmartConfig report
# author          :Thanga Prakash
# version         :1.0
############################################################


import os, sys
from sys import getsizeof
import xml.etree.ElementTree as ET
import xmltodict, json
import time
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.exceptions import InvalidFileException
from PIL import Image
from pure_dir.infra.apiresults import *
from pure_dir.components.compute.ucs.ucs_report import *
from pure_dir.components.storage.mds.mds_report import *
from pure_dir.components.storage.purestorage.pure_report import *
from pure_dir.components.network.nexus.nexus_report import *
from pure_dir.services.apps.pdt.core.orchestration.orchestration_workflows import flash_stack_type_api
from pure_dir.services.apps.pdt.core.topology import *
from pure_dir.services.apps.pdt.core.systemmanager import system_info
from pure_dir.components.generic_details import get_fs_components, get_fs_connections


report_file = '/mnt/system/pure_dir/pdt/report.xml'
topology_file = '/var/www/html/static/images/topologies/'
generated_report = '/var/www/html/static/downloads/SmartConfigReport.xlsx'
logo_path = '/var/www/html/static/images/logo.png'

class SCReport:
    def __init__(self):
        pass

    def report(self, stacktype):
        res = result()
        resObj = []
        self.alter_san_type(stacktype) 
	try:
	    # Open & read the configuration content from the report xml
            fd = open(report_file, 'r')
	    root = ET.fromstring(fd.read())

	    # Select all 'section' nodes from the report template
	    for section in root:
		# Check whether to display this section or not
		if 'stacktypes' in section.attrib and section.attrib['stacktypes'] != '':
		    if stacktype not in section.attrib['stacktypes'].split('|'):
		        continue

	        obj = {'header': {}, 'footer': {}, 'args': []}
		# Fetch arguments if any
	        for args in section.findall('./args/arg'):
		    arg = {}
		    arg['name'] = args.attrib['name']
		    arg['value'] = args.attrib['value']
		    obj['args'].append(arg)

		if 'api' in section.attrib and section.attrib['api'] != '':
                    obj['api'] = section.attrib['api']
		if 'belongsTo' in section.attrib and section.attrib['belongsTo'] != '':
                    obj['belongsTo'] = section.attrib['belongsTo']
		if 'header_title' in section.attrib and section.attrib['header_title'] != '':
                    obj['header']['title'] = section.attrib['header_title']
		if 'header_desc' in section.attrib and section.attrib['header_desc'] != '':
                    obj['header']['desc'] = section.attrib['header_desc']
		if 'footer_title' in section.attrib and section.attrib['footer_title'] != '':
                    obj['footer']['title'] = section.attrib['footer_title']
		if 'footer_desc' in section.attrib and section.attrib['footer_desc'] != '':
                    obj['footer']['desc'] = section.attrib['footer_desc']
                resObj.append(obj)
            fd.close()
            code = PTK_OKAY
            msg = 'Success'
	
	except IOError as error:
            code = PTK_INTERNALERROR
            msg = 'Could not read file: %s' % report_file
	except AttributeError as error:
            code = PTK_INTERNALERROR
            msg = 'Invalid method call occured'
        except Exception as error:
            code = PTK_INTERNALERROR
            msg = 'Unexpected eror occured. Please try again later'

        if code != PTK_OKAY:
            # Log message incase of errors
            loginfo(msg)
     
	res.setResult(resObj, code, msg)
        return res

    def report_info(self, method, args):
        res = result()
	resObj = {'labels': [], 'list': []}
        tmpObj = {}
	args = eval(json.dumps(args))
	try:
	    # Call reporting methods dynamically received from the request
	    code, data, msg = getattr(sys.modules[__name__], "%s" % method)(args['keys'])
	    # Open & read the configuration content from the report xml
            fd = open(report_file, 'r')
	    root = ET.fromstring(fd.read())

	    # Select all 'label' nodes that are children of section with attribute api='{method}' 
            for section in root.findall(".//*[@api='" + method + "']"):
                for label in section.findall(".//labels/label"):
                    tmpObj = {'key': label.attrib['key'], 'label': label.text}
                    if 'width' in label.attrib:
                        tmpObj['cellWidth'] = label.attrib['width']
                    if 'rowMerge' in label.attrib:
                        tmpObj['rowMerge'] = label.attrib['rowMerge']
                    resObj['labels'].append(tmpObj)
                break
            fd.close()
	except IOError as error:
	    data = False
	    code = PTK_INTERNALERROR
	    msg = 'Could not read file: %s' % report_file
	except AttributeError as error:
	    data = False
	    code = PTK_INTERNALERROR
	    msg = 'Invalid method call occured (%s)' % method
	except Exception as error:
	    data = False
	    code = PTK_INTERNALERROR
	    msg = 'Unexpected eror occured. Please try again later'

	if code == PTK_OKAY:
	    resObj['list'] = data
	else:
	    # Log message incase of errors
	    loginfo(msg)

        res.setResult(resObj, code, msg)
        #loginfo(resObj)
        return res

    def generate_excel(self, stacktype): 
        """ 
        Generates SmartConfig Report in .xlsx format using Openpyxl package

        Parameters:
            stacktype (str): Stacktype for which report is to be generated.
        """
        report_api_res = []
        fs_con_data = {}
        sheet_index = 0
        stack_label = "" 
        start_col_alpha = 'B'
        
        # Border styles
        cell_border = Border(left=Side(style='thin'), 
                      right=Side(style='thin'), 
                      top=Side(style='thin'), 
                      bottom=Side(style='thin'))

        # Workbook creation based on the status of Report API
        report_api_res = self.report(stacktype).getResult()
        if self.report(stacktype).getStatus() == 0:
            work_book = Workbook()
            work_book.save(generated_report)
        
        # Incase of exception in Report API
        try: 
            wb = load_workbook(generated_report)
        
        except InvalidFileException as ife:
            loginfo("An exception occured in the report API")
            loginfo("Smart config report file not generated" + str(ife))

        try:
            #Gather data from all threads
            api_results = self.initiate_sheet_thread(report_api_res)
            if api_results.get('fs_con_data') != {}:
                fs_con_data = api_results['fs_con_data']
            #Landing Sheet Creation 
            sheet = wb.create_sheet(index = sheet_index, title = "General Info")
            sheet_index += 1
            
            if "rack" in stacktype:
                stacktype = stacktype.split("-rack")[0]
    
            self.generate_generic_info(sheet, logo_path, stacktype, cell_border, start_col_alpha, fs_con_data, wb)

            # Topology diagram
            topo_img_path = topology_file+stacktype+".png"
            sheet = wb.create_sheet(index = sheet_index, title = "FlashStack Topology")
            sheet_index += 1
            self.generate_topology_info(sheet, stacktype, topo_img_path, fs_con_data, start_col_alpha, cell_border, wb)
            self.write_sheets_excel(report_api_res, api_results, wb, sheet_index, start_col_alpha, cell_border)
        except Exception as e:
            loginfo("An exception occured while writing the table " + str(e))
            pass

        finally:
            wb.save(generated_report)
	    release_ucsm_handler()
            wb.close()

    # Initiate multiple threads for the sheets in report xml
    def initiate_sheet_thread(self, report_api_res):
        pages = list({report_res['belongsTo'] for report_res in report_api_res})
        pages.append('Topology Data')
        api_dict = {}
        api_results = {}
        thread_page = {}
        for page in pages:
            #Create FSConnectivity thread along with page thread
            if page == 'Topology Data':
                api_dict[page] = []
            else:
                api_dict[page] = [report_res for report_res in report_api_res  if report_res['belongsTo'] == page]
        for page in pages:
            thread_page[str(page)] = threading.Thread(target=self.call_method_by_page, args=(api_dict[page], api_results))
            thread_page[str(page)].start()

        for page in pages:
            thread_page[str(page)].join()
        return api_results

    # Call the individual APIs in a page 
    def call_method_by_page(self, api_dict, api_results):
         if api_dict == []:
            api_results['fs_con_data'] = fs_connectivity().getResult()
         else:
            for api_list in  api_dict:
                method = api_list['api']
                args = {"keys": {}}
                if api_list['args'] != []:
                    for arg in api_list['args']:
                        args['keys'][arg['name']] = arg['value']

                #For methods with args
                if args['keys'] != {}:
                    method_res = api_list['api'] + args['keys'][arg['name']]
                else:
                    method_res = api_list['api']

                api_results[str(method_res)] = self.report_info(method, args).result

    
    # Write output from the threads into Excel
    def write_sheets_excel(self, report_api_res, api_results, wb, sheet_index, start_col_alpha, cell_border):
        for report_res in report_api_res:
            headers = []
            keys = []
            merge = False
            merge_cols = 0
            try:
                title = report_res['header']['title']
                sheet_name = report_res['belongsTo']

                if sheet_name not in wb.sheetnames:
                    sheet = wb.create_sheet(index = sheet_index, title = sheet_name)
                    sheet_index += 1
                    sheet._current_row = 2
                else:
                    sheet = wb[sheet_name]
                args = {"keys": {}}

                if report_res['args'] != []:
                    for arg in report_res['args']:
                        args['keys'][arg['name']] = arg['value']

                #For methods with args
                if args['keys'] != {}:
                   method = report_res['api'] + args['keys'][arg['name']]
                else:
                   method = report_res['api']

                report_info_output = api_results[method]
                for header_iter in range(len(report_info_output['labels'])):
                    if report_info_output['labels'][header_iter]['key'] in report_info_output['list'][0].keys():
                        keys.append(report_info_output['labels'][header_iter]['key'])
                        headers.append(report_info_output['labels'][header_iter]['label'])
                        if "rowMerge" in report_info_output['labels'][header_iter].keys():
                            merge = True
                            merge_cols += 1

                # Table Title Section
                length = len(headers)
                (cell_range, start_col, till_col) = self.generate_cell_range(length, sheet._current_row, start_col_alpha)
                self.generate_title_excel(cell_range, start_col, till_col, title, sheet, cell_border)

                # Table Header section
                (cell_range, start_col, till_col) = self.generate_cell_range(length, sheet._current_row, start_col_alpha)
                self.generate_header_excel(cell_range, start_col, till_col, headers, sheet, cell_border)

                # Table Data Section
                self.generate_data_excel(length, keys, report_info_output['list'], sheet, cell_border, start_col_alpha, merge, merge_cols)

            except Exception as e:
                loginfo("An exception occured while writing the table " + title + str(e))
                continue

            finally:
                sheet.sheet_view.showGridLines = False
                wb.save(generated_report)
                sheet._current_row += 3

    #Generate Generic Info Table
    def generate_generic_info(self, sheet, logo_path, stacktype, cell_border, start_col_alpha, fs_con_data, wb):
        try:
            #Insert Logo Images
            system_info_res = system_info().getResult()
            path_pref = "/var/www/html/static/images/"
            if os.path.exists(logo_path):
                width = 200
                height = 50
                if ((system_info_res.get('report_logo') != None) and 
                   (os.path.exists(path_pref + system_info_res.get('report_logo')))):
                    # Add default logo image to the left of the sheet
                    sheet.merge_cells('B2:D5')
                    self.add_image_sheet(sheet, 'B2', logo_path, width, height)
                    # Add uploaded logo image to the right of the sheet
                    sheet.merge_cells('F2:H5')
                    self.add_image_sheet(sheet, 'G2', path_pref + system_info_res.get('report_logo'), width, height)
                else:
                    sheet.merge_cells('E2:H5')
                    # Add default logo image to sheet
                    self.add_image_sheet(sheet, 'E2', logo_path, width, height)
                # For an image of size 208 x 52 nearly 5 rows are occupied
                sheet._current_row += 5
            else:
                loginfo("Logo image doesn't exist" + str(ioe))
                sheet._current_row += 1
                pass
            stack_label = get_stack_details()['label']
            if stacktype.split('-')[-1].lower() == 'fc':
                stack_label += " [FC]"
            else:
                stack_label += " [iSCSI]"
            
            # Write title for Generic Information
            title =  "FLASHSTACK SMARTCONFIG REPORT"
            tool_version = 'Tool Version: ' + system_info_res['version']
            self.write_generic_title(sheet, title, stack_label, tool_version)            
            general_list = get_fs_components(fs_con_data)[0]
            header_general = general_list[0]
            general_list.pop(0)

            # Table Title Section
            length = len(general_list[0].values())
            title = "GENERIC INFORMATION"
            merge = False
            merge_cols = 0
            (cell_range, start_col, till_col) = self.generate_cell_range(length, sheet._current_row, start_col_alpha)
            self.generate_title_excel(cell_range, start_col, till_col, title, sheet, cell_border)

            # Table Header section
            (cell_range, start_col, till_col) = self.generate_cell_range(length, sheet._current_row, start_col_alpha)
            self.generate_header_excel(cell_range, start_col, till_col, header_general.values(), sheet, cell_border)
 
            # Table Data Section
            self.generate_data_excel(length, header_general.keys(), general_list, sheet, cell_border, start_col_alpha, merge, merge_cols)

        except Exception as e:
            loginfo("An exception occured while writing the Generic Infomation table " +  str(e))
            pass

        finally:
            sheet.sheet_view.showGridLines = False
            wb.save(generated_report)

    #Generate Topology diagram and cabling tables 
    def generate_topology_info(self, sheet, stacktype, topo_img_path, fs_con_data, start_col_alpha, cell_border, wb):
        try:
            component_dict = {}
            device_name_dict = {}
            if os.path.exists(topo_img_path):
                #A width of 50 columns is assumed 
                sheet.merge_cells('B1:AY2')
                title_cell = sheet.cell(row=1, column=2)
                title = "TOPOLOGY DIAGRAM - SMARTCONFIG"
                title_cell.value = title
                title_cell.font = Font(size=24, bold=True)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                img = Image.open(topo_img_path)                
                img_width, img_height = img.size
                #UCSMINI stacktypes alone have resized height as 1350
                if 'ucsmini' in stacktype.split('-'):
                    res_height = 1350
                else:
                    res_height = 2000
                #Formula to calculate the resized width from resized height 
                res_width = int(img_width*(float(res_height)/img_height))
                # Add image to sheet
                self.add_image_sheet(sheet, 'B5', topo_img_path, res_width, res_height)
                # For an image resized to 1000 X 1000 nearly 60 rows are occupied
                sheet._current_row += int(math.ceil((58.0/1000)*res_height))
            else:
                loginfo("Topology image doesn't exist")
                sheet._current_row += 2
            
            if fs_con_data.get('connections') != {}:
                component_dict = get_fs_components(fs_con_data)[1]
                device_name_dict = get_fs_components(fs_con_data)[2]
            component_dict = {k:sorted(v) for k,v in component_dict.items()}
            device_name_dict = dict(sorted(device_name_dict.items()))
            merge = True
            merge_cols = 1
            for components in component_dict.values():
                for component in components:
                    try:
                        fs_result = []
                        header_topo = {}
                        fs_result=get_fs_connections(fs_con_data,component, device_name_dict[component])
                        header_topo = fs_result[0]
                        fs_result.pop(0)

                        # Table Title Section
                        length = len(header_topo.values())
                        title = component + " Cabling Information"
                        (cell_range, start_col, till_col) = self.generate_cell_range(length, sheet._current_row, start_col_alpha)
                        self.generate_title_excel(cell_range, start_col, till_col, title, sheet, cell_border)
                        # Table Header section
                        (cell_range, start_col, till_col) = self.generate_cell_range(length, sheet._current_row, start_col_alpha)
                        self.generate_header_excel(cell_range, start_col, till_col, header_topo.values(), sheet, cell_border)
                        # Table Data Section
                        self.generate_data_excel(length, header_topo.keys(), fs_result, sheet, cell_border, start_col_alpha, merge, merge_cols)
                    except Exception as e:
                        loginfo("An exception occured while writing the Cabling Information " + str(e))
                        continue
                    finally:
                        sheet._current_row += 3
                        wb.save(generated_report)
        except Exception as e:
            loginfo("An exception occured while writing Topology and Cabling Information" + str(e))
            pass
        finally:
            sheet.sheet_view.showGridLines = False
            wb.save(generated_report)

    # Based on the row count returns the cell range to be written to Excel
    def generate_cell_range(self, length, row_count, start_col_alpha):
        cell_range = []
        if (ord(start_col_alpha) + length) <= ord('Z'):
            for cell_count in range(length):
                cell_range.append(chr(ord(start_col_alpha) + cell_count) + str(row_count))
        else:
            for cell_count in range((ord('Z') - ord(start_col_alpha)) + 1):
                cell_range.append(chr(ord(start_col_alpha) + cell_count) + str(row_count))
            for cell_count in range(length  - ((ord('Z') - ord(start_col_alpha)) + 1)):
                cell_range.append('A' + chr(ord('A') + cell_count) + str(row_count))
        start_col = cell_range[0]
        till_col = cell_range[-1]
        return cell_range, start_col, till_col

    # Generate Title Section in Excel
    def generate_title_excel(self, cell_range, start_col, till_col, title, sheet, cell_border):
        for title_count in range(len(cell_range)):
            sheet[cell_range[title_count]].border = cell_border
            sheet.merge_cells(start_col + ":" + till_col)
            title_cell = sheet.cell(row=sheet._current_row, column=(ord(start_col[0])-ord('A')) + 1)
            title_cell.value = title
            title_cell.fill = PatternFill("solid", fgColor="FB5000")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            title_cell.font = Font(size = 14, bold = True, color = 'FFFFFF')
        sheet._current_row += 1

    # Generate Header Section in Excel
    def generate_header_excel(self, cell_range, start_col, till_col, headers, sheet, cell_border):
        for header_count in range(len(cell_range)):
            sheet[cell_range[header_count]].fill = PatternFill("solid", fgColor='FBA882')
            sheet[cell_range[header_count]].alignment = Alignment(horizontal='center', vertical='center')
            sheet[cell_range[header_count]].value = headers[header_count]
            sheet[cell_range[header_count]].font = Font(size = 12, bold = True, color = 'FFFFFF')
            sheet[cell_range[header_count]].border = cell_border
        sheet._current_row += 1

    # Generate Data Section in Excel
    def generate_data_excel(self, length, keys, data_list, sheet, cell_border, start_col_alpha, merge, merge_cols):
        start_row = str(sheet._current_row)
        for data in data_list:
            for key,val in data.items():
                if isinstance(val, list):
                    data[key] = "\n".join(val)
                elif isinstance(val, dict):
                    dict_data = ""
                    for k,v in val.items():
                        dict_data += v['label'] + " : " + v['value'] + "\n"
                    data[key] = dict_data
            (cell_range, start_col, till_col) = self.generate_cell_range(length, sheet._current_row, start_col_alpha)
            for iter_cnt in range(len(cell_range)):
                sheet[cell_range[iter_cnt]].value = data[keys[iter_cnt]]
                sheet[cell_range[iter_cnt]].border = cell_border
                sheet[cell_range[iter_cnt]].alignment = Alignment(horizontal='left', vertical='top')
            sheet._current_row += 1
        end_row = str(sheet._current_row-1)

        if merge==True:
            for col_count in range(0,merge_cols):
                sheet.merge_cells(chr(ord(start_col_alpha) + col_count) + start_row + ':' + chr(ord(start_col_alpha) + col_count) + end_row)

        for column_cells in sheet.columns:
            unmerged_cells = list(filter(lambda cell_to_check: cell_to_check.coordinate not in sheet.merged_cells, column_cells))
            length_list = []
            for cell in unmerged_cells:
                string_list = []
                if "\n" in str(cell.value):
                    string_list = str(cell.value).split("\n")
                    length_list.append(max(len(string) for string in string_list))
                else:
                    length_list.append(len(str(cell.value)))
            length = max(length_list)
            for cell in unmerged_cells:
                if "\n" in str(cell.value):
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
                else:
                    if len(str(cell.value)) > length:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrapText=True)
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='top')
            sheet.column_dimensions[unmerged_cells[0].column_letter].width = length

    # Write title for Generic Info table
    def write_generic_title(self, sheet, title, stack_label, tool_version):
        sheet.merge_cells('A' + str(sheet._current_row) + ':K' + str(sheet._current_row + 2))
        title_cell = sheet.cell(row=sheet._current_row, column=1)
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText = True)
        title_cell.font = Font(size = 32, bold = True)
        sheet._current_row += 3
        sheet.merge_cells('A' + str(sheet._current_row) + ':K' + str(sheet._current_row + 1))
        stack_cell = sheet.cell(row=sheet._current_row, column=1)
        stack_cell.value = tool_version
        stack_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText = True)
        stack_cell.font = Font(size =18, bold = True)
        sheet._current_row += 3
        sheet.merge_cells('A' + str(sheet._current_row) + ':K' + str(sheet._current_row + 2))
        stack_cell = sheet.cell(row=sheet._current_row, column=1)
        stack_cell.value = stack_label
        stack_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText = True)
        stack_cell.font = Font(size = 24, bold = True)
        sheet._current_row += 4
      
    # Add image to sheet 
    def add_image_sheet(self, sheet, img_cell, img_path, width, height):
        Image.MAX_IMAGE_PIXELS = None
        img = Image.open(img_path)
        img = img.resize((width,height),Image.LANCZOS)
        img.save(img_path)
        new_img = OpenpyxlImage(img_path)
        sheet.add_image(new_img, img_cell)
  
    def generate_report(self, stacktype):
        '''
        executes generate excel using threading
        inputs: stacktype
        returns: created thread id and status
        '''
        res = result()
        resObj = {'tid':'', 'report_status':''}
        if os.path.exists(generated_report):
            loginfo("SmartConfig report already exists")
            resObj['tid'] = generated_report.split("/")[-1]
            resObj['report_status'] = 'Completed'
            code = PTK_OKAY
            msg = "Success"
            res.setResult(resObj, code, msg)
            return res
        else:
            try:
                self.alter_san_type(stacktype)
                thread = threading.Thread(target=self.generate_excel, args=(stacktype,))
                thread.start()
                loginfo("Thread has been initiated to generate report")
                thread.join()
                resObj['tid']= str(thread.ident)
                resObj['report_status'] = 'started'
                code = PTK_BACKGROUND_OPER
                msg = "Success"
                res.setResult(resObj, code, msg)
                return res
            except Exception as e:
                loginfo("Unexcepted Error" +str(e))
                code = PTK_INTERNALERROR
                msg = "Failed"
                res.setResult(resObj, code, msg)
                return res
    
    def report_status(self, tid):
        '''
        inputs: thread id
        returns: thread status and status of report
        '''
        res = result()
        thread_list = [str(i) for i in threading.enumerate() if tid in str(i)]
        resObj = {'tid':'', 'report_status':''}
        if thread_list != []:
            loginfo("Report is being generated........")
            resObj['tid'] = tid
            resObj['report_status']='In-Progress'
            code = PTK_BACKGROUND_OPER
            msg = "Success"
            res.setResult(resObj, code, msg) 
            return res
        else:
            if os.path.exists(generated_report):
                loginfo("Report has been generated successfully")
                resObj['tid'] = generated_report.split("/")[-1]
                resObj['report_status'] = 'Completed'
                code = PTK_OKAY
                msg = "Success"
                res.setResult(resObj, code, msg)
                time.sleep(10)
                return res
            else:
                loginfo("ReportGeneration failed")
                code = PTK_INTERNALERROR
                msg = "Failed"
                res.setResult(resObj, code, msg)
                return res
       
    def alter_san_type(self, stacktype):
        """
        Writes SAN Type arg in the report config file based on stacktype.
        """
        san_type = ""
        if "rack" in stacktype:
            stacktype = stacktype.split("-rack")[0]

        if stacktype.upper().find("FC") != -1:
            san_type = "FC"
        elif stacktype.upper().find("ISCSI") != -1:
            san_type = "iSCSI"

        with open(report_file) as r_file:
            rf_doc = xmltodict.parse(r_file.read())
            for section in rf_doc['sections']['section']:
                if 'args' in section.keys():
                    if section['args']['arg']['@name'] == 'san_type':
                        section['args']['arg']['@value'] = san_type

        rf_out = xmltodict.unparse(rf_doc, pretty=True)
        with open(report_file,'w') as w_file:
            w_file.write(rf_out.encode('utf-8'))

