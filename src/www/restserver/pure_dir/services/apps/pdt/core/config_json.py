#!/usr/bin/env python
# Project_Name    :FlashStack SmartConfig
# title           :config_json.py
# description     :Export And Import Configuration
# version         :1.0
###################################################################

import threading
from xml.dom.minidom import *
import shutil
import shelve
import os
import time
import xmltodict
import json
import copy
from collections import OrderedDict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.styles.borders import Border, Side

from pure_dir.infra.logging.logmanager import loginfo
from pure_dir.infra.apiresults import *

from pure_dir.services.apps.pdt.core.discovery import*
from pure_dir.services.apps.pdt.core.orchestration.orchestration_config import*
from pure_dir.services.apps.pdt.core.orchestration.orchestration_workflows import g_flash_stack_types
from pure_dir.services.apps.pdt.core.orchestration.orchestration import*
from pure_dir.services.apps.pdt.core.tasks.main.ucs import*
from pure_dir.services.apps.pdt.core.tasks.main.pure import *
from pure_dir.services.apps.pdt.core.tasks.main.nexus_5k import*
from pure_dir.services.apps.pdt.core.tasks.main.nexus_9k import*
from pure_dir.services.apps.pdt.core.tasks.main.mds import*
from pure_dir.services.apps.pdt.core.systemmanager import *
from pure_dir.global_config import get_settings_file, get_discovery_store
import glob


def get_wfflag_dicts(stacktype):
    '''Returns a dict of alternative workflow id's in case of workflow flags being True.'''
    flag_dict = {}
    flag_list = get_wfflags()
    files = "/mnt/system/pure_dir/pdt/workflows/" + stacktype + "/*.xml"
    wf_files = glob.glob(files)
    for wf_file in wf_files:
        with open(wf_file) as td:
            wf_doc = xmltodict.parse(td.read())
            if wf_doc['workflow'].get('wfs'):
                for wf in getAsList(wf_doc['workflow']['wfs']['wf']):
                    if wf.get('@flags'):
                        for flag in wf['@flags'].split(";"):
                            if flag.split(":")[0] in flag_list:
                                flag_dict[wf['@id']] = flag.split(":")[1]
    return flag_dict


def get_wfflags():
    '''Returns a list of flags if True.'''
    flag_list = []
    if os.path.exists(get_skip_flag()):
        with open(get_skip_flag()) as td:
            wf_doc = xmltodict.parse(td.read())
            for flag in getAsList(wf_doc['wflags']['wflag']):
                if flag['@flag'] == 'True':
                    flag_list.append(flag['@name'])
    return flag_list


def export_configuration(stacktype, path=None):
    res = result()
    dicts = {}
    grp_wf_list = []
    flag_alter_wid = get_wfflag_dicts(stacktype)
    os.chdir('/mnt/system/pure_dir/pdt/sreq/')
    for shlv_file in glob.glob("*.shlv"):
        fl = shlv_file.split('.')[0][4:]
        shelf = shelve.open(get_shelf_file(fl), flag='r')
        wflist = []
        for x in range(len(shelf['job'])):
            job_recs = shelf['job'][str(x)]['record']
            job_rec = job_recs['tasklist']
            keylist = ['taskstatus', 'inputs', 'outputs', 'class']
            wfdict = {
                "wfname": job_recs['wfname'], "desc": job_recs['desc'], "wid": job_recs['wid']}
            tasklist = []

            job_doc = None
            if job_recs['jobid'] is not None and os.path.exists(
                    get_job_file(job_recs['jobid'])) is True:
                job_file = get_job_file(job_recs['jobid'])
                fd = open(job_file, 'r')
                job_doc = xmltodict.parse(fd.read())

            wf_doc = None

            # Check altenative wid in case of workflow flag
            if job_recs['wid'] in flag_alter_wid.keys():
                fname = get_workflow_file(flag_alter_wid[job_recs['wid']], stacktype)
            else:
                fname = get_workflow_file(job_recs['wid'], stacktype)

            with open(fname) as td:
                wf_doc = xmltodict.parse(td.read())

            for key, value in job_rec.items():
                data_list = []
                for k, v in value['inputs'].items():
                    data = get_value_from_xml_updated(
                        job_doc, value['texecid'], k)
                    s = {
                        "name": k,
                        "value": data[0],
                        "mapval": data[1]
                    }
                    wf_data = get_value_from_xml_updated(
                        wf_doc, value['texecid'], k)

                    if s["value"] != wf_data[0] or s["mapval"] != wf_data[1]:
                        if wf_data[2]:   # Check for ovewrite Flag
                            data_list.append(s)

                value['input'] = data_list
                for ky in keylist:
                    if ky in value:
                        value.pop(ky)
                if data_list != []:
                    tasklist.append(value)
            wfdict['tasks'] = tasklist

            if wfdict['tasks'] != []:
                wflist.append(wfdict)
        grp_wf = {
            "grp_workflow_name": shelf['name'],
            "grp_workflow_id": shelf['wid'],
            "subworkflows": wflist
        }
        if wflist != []:
            grp_wf_list.append(grp_wf)

    ucs_kvm_console_ip = ''
    if os.path.exists(get_global_wf_config_file()):
        doc = parse(get_global_wf_config_file())
        htype_list = doc.getElementsByTagName('htype')
        input_list = []
        #input_list.append({'README':"feel free to update 'value' field under global_config session"})
        for htype in htype_list:
            if htype.getAttribute('stacktype') == stacktype:
                for ipt in htype.getElementsByTagName('input'):
                    if ipt.getAttribute('name') in [
                        'nexus_switch_a',
                        'nexus_switch_b',
                        'mds_switch_a',
                        'mds_switch_b',
                        'ucs_switch_a',
                        'ucs_switch_b',
                        'pure_id',
                            'remote_file']:
                        # remote_file to be taken from UCSM
                        continue

                    if ipt.getAttribute('name') == 'kvm_console_ip':
                        ucs_kvm_console_ip = ipt.getAttribute('value')
                    else:
                        input_dict = {
                            "name": ipt.getAttribute('name'),
                            "value": ipt.getAttribute('value'),
                            "description": ipt.getAttribute('helptext')
                        }
                        input_list.append(input_dict)
    discovery_list = []
    if os.path.exists(get_devices_wf_config_file()):
        doc = parse(get_devices_wf_config_file())
        device_list = doc.getElementsByTagName('device')
        for device in device_list:
            discovery_dict = {}
            discovery_dict["name"] = device.getAttribute('name')
            discovery_dict["device_type"] = device.getAttribute('device_type')
            if device.getAttribute('device_type') not in ['PURE', 'FlashBlade']:
                discovery_dict["tag"] = device.getAttribute('tag')
                discovery_dict["switch_ip"] = device.getAttribute('ipaddress')

            if device.getAttribute('device_type') == 'Nexus 5k':
                discovery_dict['system_image'] = device.getAttribute(
                    'switch_system_image')
                discovery_dict['kickstart_image'] = device.getAttribute(
                    'switch_kickstart_image')
                discovery_dict['image_version'] = device.getAttribute(
                    'image_version')
            elif device.getAttribute('device_type') == 'Nexus 9k':
                discovery_dict['system_image'] = device.getAttribute(
                    'switch_image')
                discovery_dict['image_version'] = device.getAttribute(
                    'image_version')
            elif device.getAttribute('device_type') == 'MDS':
                discovery_dict['kickstart_image'] = device.getAttribute(
                    'switch_kickstart_image')
                discovery_dict['system_image'] = device.getAttribute(
                    'switch_system_image')
                discovery_dict['image_version'] = device.getAttribute(
                    'image_version')
            elif device.getAttribute('device_type') == 'UCSM':
                discovery_dict['leadership'] = device.getAttribute(
                    'leadership')
                discovery_dict['virtual_ip'] = device.getAttribute(
                    'vipaddress')
                # if  discovery_dict['leadership'] ==
                if discovery_dict['leadership'] == 'primary':
                    discovery_dict['kvm_console_ip'] = ucs_kvm_console_ip
                    discovery_dict['esxi_file'] = device.getAttribute(
                        'esxi_file')
                    discovery_dict['esxi_kickstart'] = device.getAttribute(
                        'esxi_kickstart')

                discovery_dict['dns'] = device.getAttribute('dns')
                discovery_dict['domain_name'] = device.getAttribute(
                    'domain_name')
                discovery_dict['blade_image'] = device.getAttribute(
                    'blade_image')
                discovery_dict['infra_image'] = device.getAttribute(
                    'infra_image')
                discovery_dict['server_type'] = device.getAttribute(
                    'server_type')
            elif device.getAttribute('device_type') in ['PURE', 'FlashBlade']:
                if device.hasAttribute('isZTP'):
                    discovery_dict['ZTP'] = True
                    discovery_dict['ct0_ip'] = device.getAttribute('ct0_ip')
                    discovery_dict['ct1_ip'] = device.getAttribute('ct1_ip')
                    discovery_dict['vir0_ip'] = device.getAttribute('vir0_ip')
                    discovery_dict['domain_name'] = device.getAttribute('domain_name')
                    discovery_dict['relay_host'] = device.getAttribute('relay_host')
                    discovery_dict['sender_domain'] = device.getAttribute('sender_domain')
                    discovery_dict['dns'] = device.getAttribute('dns')
                    discovery_dict['alert_emails'] = device.getAttribute('alert_emails')
                    discovery_dict['organization'] = device.getAttribute('organization')
                    discovery_dict['full_name'] = device.getAttribute('full_name')
                    discovery_dict['job_title'] = device.getAttribute('job_title')
                    discovery_dict['timezone'] = device.getAttribute('timezone')
                else:
                    discovery_dict['array_ip'] = device.getAttribute('ipaddress')

            if bool(discovery_dict):
                discovery_list.append(discovery_dict)
    dicts['components'] = discovery_list
    dicts['workflows'] = grp_wf_list
    dicts['global_config'] = input_list
    dicts['stacktype'] = stacktype
    dicts['version'] = '1.0'
    dicts['tool_version'] = get_smartconfig_version()
    dicts['flags'] = ':'.join(get_wfflags())

    if path is None:
        prettfy_json(g_download_dir + stacktype + '.json', dicts)
    else:
        prettfy_json(path + stacktype + '.json', dicts)

    res.setResult({"url": stacktype + ".json"}, PTK_OKAY, "success")
    return res


def get_component_description(tag, device_type, field):
    read_only_str = "[type: read-only]"
    editable_str = "[type: editable]"
    if field == "dns":
        return "Name server IP: DNS server IP(s) (Example: 10.1.164.9) " + editable_str
    if field == "domain_name":
        return "DNS domain name (Example: flashstack.cisco.com)" + editable_str
    if field == "esxi_file":
        return "VMware vSphere ESXi Cisco Custom ISO (Eg:VMware_ESXi_6.7.0_10302608_Custom_Cisco_6.7.1.1) " + editable_str
    if field == 'leadership':
        return "leadership primary/subordinate " + editable_str
    if field == 'name':
        if device_type == 'UCSM' and tag == 'A':
            return "Cisco UCS Fabric Interconnect Name (Example:fi-6332-A) " + editable_str
        if device_type == 'UCSM' and tag == 'B':
            return "Cisco UCS Fabric Interconnect Name (Example:fi-6332-B) " + editable_str
        if device_type == 'MDS' and tag == 'A':
            return "Cisco MDS hostname (Example: mds-9132T-A) " + editable_str
        if device_type == 'MDS' and tag == 'B':
            return "Cisco MDS hostname (Example: mds-9132T-B) " + editable_str
        if device_type == 'Nexus 9k' and tag == 'A':
            return "Nexus-A name" + editable_str
        if device_type == 'Nexus 9k' and tag == 'B':
            return "Nexus-B name" + editable_str
        if device_type == 'Nexus 5k' and tag == 'A':
            return "Nexus-A name" + editable_str
        if device_type == 'Nexus 5k' and tag == 'B':
            return "Nexus-B name" + editable_str
        if device_type == 'FlashArray':
            return "FlashArray name " + editable_str
        if device_type == 'FlashBlade':
            return "FlashBlade name " + editable_str

        return "name " + editable_str
    if field == 'switch_ip':
        return device_type + " Management IP address " + editable_str
    if field == 'virtual_ip':
        return device_type + " Virtual IP Address " + editable_str
    if field == 'kvm_console_ip':
        return "KVM Console IP Address Range" + editable_str
    if field == 'array_ip':
        return device_type + " IP Address " + editable_str
    if field == 'server_type':
        return "UCS server type Blade/Rack " + read_only_str
    if field == 'blade_image':
        return "UCS Blade firmware image " + editable_str
    if field == 'infra_image':
        return "UCS Infrastructure Software Bundle " + read_only_str
    if field == 'system_image':
        if device_type == 'Nexus 9k' and tag == 'A':
            return "Nexus 9k System Image for Nexus-A and Nexus-B " + editable_str

        if device_type == 'Nexus 5k' and tag == 'A':
            return "Nexus 5k System Image for Nexus-A and Nexus-B" + editable_str
        if device_type == 'MDS' and tag == 'A':
            return 'MDS System Image for MDS-A and MDS-B'

        return device_type + " System Image " + read_only_str
    if field == 'image_version':
        return device_type + " Image version " + read_only_str
    if field == 'kickstart_image':
        if device_type == 'MDS' and tag == 'A':
            return "MDS kickstart image for MDS-A " + editable_str
        return device_type + " kickstart image " + read_only_str

    if field == 'esxi_kickstart':
        return "Option to perform automated installation of ESX, specify kickstart filename " + editable_str

    if field == 'ct0_ip':
        return "FlashArray CT0 IP Address " + editable_str
    if field == 'ct1_ip':
        return "FlashArray CT1 IP Address " + editable_str
    if field == 'vir0_ip':
        return "FlashArray vir0 IP Address " + editable_str
    if field == 'alert_emails':
        return "Alert Emails " + editable_str
    if field == 'full_name':
        return "Full Name " + editable_str
    if field == 'job_title':
        return "Job Title " + editable_str
    if field == 'organization':
        return "Organization " + editable_str
    if field == 'relay_host':
        return "Relay Host " + editable_str
    if field == 'sender_domain':
        return "Sender Domain " + editable_str
    if field == 'timezone':
        return "Timezone " + editable_str
    return " " + read_only_str


def get_global_description(field, desr):
    if field == 'native_vlan':
        return "VLAN for untagged frames (Example: 2) supports any value up to 3967."
    if field == 'blade_image':
        return "Use B-Series bundle for Blade servers and C-Series bundle for Rack Servers (Example: ucs-k9-bundle-b-series.4.1.1b.B.bin)"
    if field == 'ntp':
        return "NTP IP Address (Example: 192.168.1.1) supports any IPv4 value up to 255.255.255.255"
    if field == 'mgmt_vlan':
        return "In-band management network VLAN ID (Example: 115) supports any value up to 3967"
    if field == 'native_vlan':
        return "VLAN for untagged frames (Example: 2) supports any value up to 3967"
    if field == 'vmotion_vlan':
        return "VLAN for vMotion (Example: 200) supports any value up to 3967"
    if field == 'app1_vlan':
        return "Application network VLAN ID (Example: 201) supports any value up to 3967"
    if field == 'app2_vlan':
        return "Application network VLAN ID (Example: 202) supports any value up to 3967"
    if field == 'app3_vlan':
        return "Application network VLAN ID (Example: 203) supports any value up to 3967"
    if field == 'up_port_a':
        return 'Uplink Port-Channel for Fabric A between NEXUS and FI (Example: 151)'
    if field == 'up_port_b':
        return 'Uplink Port-Channel for Fabric B between NEXUS and FI (Example: 152)'
    if field == 'gateway':
        return 'Out-of-band management network gateway (Example: 192.168.164.254)'
    if field == 'firmware':
        return "Use B-Series bundle for Blade servers and C-Series bundle for Rack Servers"
    if field == 'upgrade':
        return "Option to perform UCS Infra and Server firmware upgrade Yes/No"
    if field == 'netmask':
        return 'Out-of-band management network netmask (Example: 255.255.255.0)'
    if field == 'ntp_vlan':
        return 'VLAN for in-band management interfaces (Example: 215) supports any value upto 3967'
    if field == 'fc_portchannel_a':
        return 'SAN Port-Channel for Fabric A between MDS and FI (Example: 1)supports any value between 1 and 256'
    if field == 'fc_portchannel_b':
        return 'SAN Port-Channel for Fabric B between MDS and FI (Example: 2)supports any value between 1 and 256'
    if field == 'san_a':
        return 'VSAN used for Fabric A between FlashArray and NEXUS and FI (Example: 101) supports any value between 1 and 4096'
    if field == 'san_b':
        return 'VSAN used for Fabric B between FlashArray and NEXUS and FI (Example: 102) supports any value between 1 and 4096'
    if field == 'host':
        return 'Email Relay Server IP Address supports any IPv4 value value upto 255.255.255.255'
    return desr


def prettfy_json(path, import_dict):
    tmp_dict = copy.deepcopy(import_dict)
    tmp_dict['global_config'] = {}
    tmp_dict['components'] = {}
    tmp_dict['system_info'] = {}
    tmp_dict['system_info']['stacktype'] = {
        'value': tmp_dict['stacktype'],
        'description': "FlashStack deployment type [type: read-only]"}
    tmp_dict['system_info']['version'] = {
        'value': tmp_dict['version'],
        'description': "Document version [type: read-only]"}
    tmp_dict['system_info']['flags'] = {
        'value': tmp_dict['flags'],
        'description': "flags [type: read-only]"}
    tmp_dict['system_info']['tool_version'] = {
        'value': tmp_dict['tool_version'],
        'description': "Generated Tool version [type: read-only]"}
    rm_list = ['tool_version', 'stacktype', 'version', 'flags']
    for rm_item in rm_list:
        del tmp_dict[rm_item]

    for global_config in import_dict['global_config']:
        tmp_dict['global_config'][global_config['name']] = {'value': global_config['value'], 'description': get_global_description(
            global_config['name'], global_config['description']) + " [type: editable]"}

    tmp_dict['workflows'] = import_dict['workflows']

    for component in import_dict['components']:
        if component['device_type'] == "PURE":
            component['device_type'] = 'FlashArray'
        comp_name = component['device_type'] + "_" + \
            component['tag'] if component.get('tag', None) else component['device_type']
        tmp_dict['components'][comp_name] = {}
        for k, v in component.items():
            if k == 'tag' or k == 'device_type':
                continue
            tmp_dict['components'][comp_name][k] = {
                'value': v, 'description': get_component_description(
                    component.get(
                        'tag', None), component['device_type'], k)}

    # add to global
    items = ['domain_name', 'dns', 'server_type', 'kvm_console_ip', 'esxi_file', 'esxi_kickstart']
    for item in items:
        tmp_dict['global_config'][item] = {
            'value': tmp_dict['components']['UCSM_A'][item]['value'],
            'description': tmp_dict['components']['UCSM_A'][item]['description']}

        del tmp_dict['components']['UCSM_A'][item]
        if item not in ['kvm_console_ip', 'esxi_file', 'esxi_kickstart']:
            del tmp_dict['components']['UCSM_B'][item]
    # remove reduntant values

    items = ['virtual_ip', 'leadership', 'infra_image']
    for item in items:
        del tmp_dict['components']['UCSM_B'][item]

    # remove unwanted values

    del tmp_dict['components']['UCSM_B']['blade_image']
    del tmp_dict['components']['UCSM_A']['blade_image']

    if 'MDS_A' in tmp_dict['components'].keys():
        del tmp_dict['components']['MDS_A']['image_version']

    if 'MDS_B' in tmp_dict['components'].keys():
        del tmp_dict['components']['MDS_B']['image_version']
        del tmp_dict['components']['MDS_B']['system_image']
        del tmp_dict['components']['MDS_B']['kickstart_image']

    if 'Nexus 9k_A' in tmp_dict['components'].keys():
        del tmp_dict['components']['Nexus 9k_A']['image_version']

    if 'Nexus 9k_B' in tmp_dict['components'].keys():
        del tmp_dict['components']['Nexus 9k_B']['image_version']
        del tmp_dict['components']['Nexus 9k_B']['system_image']

    if 'Nexus 5k_A' in tmp_dict['components'].keys():
        del tmp_dict['components']['Nexus 5k_A']['image_version']

    if 'Nexus 5k_B' in tmp_dict['components'].keys():
        del tmp_dict['components']['Nexus 5k_B']['image_version']
        del tmp_dict['components']['Nexus 5k_B']['system_image']
        del tmp_dict['components']['Nexus 5k_B']['kickstart_image']

    with open(path, 'w') as outfile:
        outfile.write(
            "#README: Please do not edit fields, which are marked as [type: readonly] in description, Also please do not edit workflow section\n")
        json.dump(OrderedDict(tmp_dict), outfile, indent=4, sort_keys=False)


def undo_json_pretty(import_dict):
    tmp_dict = copy.deepcopy(import_dict)
    tmp_dict['stacktype'] = tmp_dict['system_info']['stacktype']['value']
    tmp_dict['version'] = tmp_dict['system_info']['version']['value']
    tmp_dict['flags'] = tmp_dict['system_info']['flags']['value']
    tmp_dict['global_config'] = []
    tmp_dict['components'] = []
    items = [
        'domain_name',
        'dns',
        'blade_image',
        'server_type',
        'kvm_console_ip',
        'esxi_file',
        'esxi_kickstart']

    for k, v in import_dict['global_config'].items():
        if k not in items:
            tmp_dict['global_config'].append({'name': k, 'value': v['value']})

    for k, v in import_dict['components'].items():
        tmp_comp = {}
        for attr_k, attr_v in v.items():
            tmp_comp[attr_k] = attr_v["value"]
        if k not in ['FlashArray', 'FlashBlade'] :
            tmp_comp['tag'] = k.split("_")[1]
            tmp_comp['device_type'] = k.split("_")[0]
        else:
            tmp_comp['device_type'] = 'PURE' if k == 'FlashArray' else k
        if k in ['UCSM_A', 'UCSM_B']:
            tmp_comp['blade_image'] = import_dict['global_config']['firmware']['value']
            for item in items:
                for key, val in import_dict['global_config'].items():
                    if key == item:
                        tmp_comp[item] = val['value']
        if k == 'UCSM_B':
            tmp_comp['virtual_ip'] = import_dict['components']['UCSM_A']['virtual_ip']['value']
            if import_dict['components']['UCSM_A']['leadership']['value'] == 'primary':
                tmp_comp['leadership'] = 'subordinate'
            else:
                tmp_comp['leadership'] = 'primary'

            tmp_comp['infra_image'] = import_dict['components']['UCSM_A']['infra_image']['value']

        if k == 'MDS_B':
            tmp_comp['system_image'] = import_dict['components']['MDS_A']['system_image']['value']
            tmp_comp['kickstart_image'] = import_dict['components']['MDS_A']['kickstart_image']['value']

        if k == 'Nexus 9k_B':
            tmp_comp['system_image'] = import_dict['components']['Nexus 9k_A']['system_image']['value']

        if k == 'Nexus 5k_B':
            #del tmp_dict['components']['Nexus 5k_B']['image_version']
            tmp_comp['system_image'] = import_dict['components']['Nexus 5k_A']['system_image']['value']
            tmp_comp['kickstart_image'] = import_dict['components']['Nexus 5k_A']['kickstart_image']['value']
        tmp_dict['components'].append(tmp_comp)

    return tmp_dict


def skip_comments(file_path):
    with open(file_path) as f:
        data = ""
        for curline in f:
            if curline.startswith("#"):
                continue
            data += curline
        return json.loads(data)


def get_value_from_xml_updated(doc, texecid, name):
    if doc is not None:
        if doc['workflow']['@id'] is not None and 'wtype' not in doc['workflow']:
            for i in doc['workflow']['tasks']['task']:
                if i['@texecid'] == texecid:
                    overwrite = True
                    if isinstance(i['args']['arg'], (list,)):
                        for j in i['args']['arg']:
                            if j is not None and j['@name'] == name:
                                if '@overwrite' in j:
                                    overwrite = False if j['@overwrite'] == 'False' else True
                                return j['@value'], j['@mapval'] if '@mapval' in j else '', overwrite
                    else:
                        if i['args']['arg']['@name'] == name:
                            if '@overwrite' in i['args']['arg']:
                                overwrite = False if j['@overwrite'] == 'False' else True
                            return i['args']['arg']['@value'], i['args']['arg']['@mapval'] if '@mapval' in i['args']['arg'] else '', overwrite
    return "", "", False


def get_obj(path,param = None):
  if param != None:
      exec("%s = %s.%s" %
         ("res", 'param', path))
      return (locals()['res'])
  exec("%s = %s" %
         ("res", path))
  return (locals()['res'])

def import_configuration(configfile):
    res = result()
    res_data = {}
    excel_import = False
    configfile.save("/tmp/" + configfile.filename)
    loginfo("File:" + "/tmp/" + configfile.filename)
    tmp_data = None
    if str(configfile.filename).endswith('xlsx'):
        status, details = get_xml_element(get_settings_file(), 'subtype')
        try:
            tmp_data = excel_to_json(details[0]['stacktype'], "/tmp/" + configfile.filename)
        except Exception as e:
            res.setResult(res_data, PTK_INTERNALERROR, "Unexpected Error")
            return res
        excel_import = True
    else:
        tmp_data = skip_comments("/tmp/" + configfile.filename)

    if "system_info" not in tmp_data:
        res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_VERSION"))
        return res

    data = undo_json_pretty(tmp_data)
    if "version" not in data:
        res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_VERSION"))
        return res

    if "rack" in data['stacktype']:
        for comp in data['components']:
            if comp['device_type'] == "UCSM" and comp['server_type'] == "Rack":
                deployment_settings({'subtype': data['stacktype']})

    status, details = get_xml_element(get_settings_file(), 'subtype')

    if not excel_import:
        if details[0]['subtype'] != data['stacktype']:
            res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_STACKTYPE"))
            return res
    else:
        if details[0]['stacktype'] != data['stacktype']:
            res.setResult(res_data, PTK_INTERNALERROR, "Invalid Excel")
            return res

    if not excel_import:
        shutil.copy2("/tmp/" + configfile.filename, get_download_path() +
                     'import-' + data['stacktype'] + '.json')
    else:

        with open(get_download_path() + 'import-' + details[0]['subtype'] + '.json', 'w') as outfile:
            json.dump(tmp_data, outfile, indent=4)

    for wfs in data['workflows']:
        for wf in wfs['subworkflows']:
            if len(wf['tasks']) > 0:
                for i in wf['tasks']:
                    #exec("%s = %s" % ("input_obj",
                    #                  i['taskid'] + "." + i['taskid'] + "Inputs" + "()"))
                    input_obj = get_obj(i['taskid'] + "." + i['taskid'] + "Inputs" + "()")
                    inputs = [x for x in dir(input_obj) if not x.startswith(
                        '__') and not x.endswith('__')]
                    for j in i['input']:
                        if j['name'] not in inputs:
                            loginfo("Invalid task input field '%s' in json" % j['name'])
                            res.setResult(
                                res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_JSON"))
                            return res

    set_wf_flag(data['flags'].split(':'))
    if os.path.exists(get_skip_flag()):
        with open(get_skip_flag()) as td:
            wf_flags = xmltodict.parse(td.read())

    global_file = open(get_global_wf_config_file(), 'r')
    doc = xmltodict.parse(global_file.read())
    stacktype = False
    for k in doc['globals']['htype']:
        if k['@stacktype'] == data['stacktype']:
            stacktype = True
            for gl in data['global_config']:
                if gl['name'] == 'remote_file':
                    continue
                inpt = False
                for inp in k['input']:
                    if inp['@name'] == 'remote_file':
                        continue
                    if gl['name'] == inp['@name']:
                        inpt = True
                        break
                if not inpt:
                    res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_JSON"))
                    return res
    for comp in data['components']:
        if 'device_type' in comp and comp['device_type'] == 'Nexus 5k':
            if 'system_image' not in comp or 'kickstart_image' not in comp:
                res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_JSON"))
                return res
        elif 'device_type' in comp and comp['device_type'] == 'Nexus 9k':
            if 'system_image' not in comp:

                res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_JSON"))
                return res
        elif 'device_type' in comp and comp['device_type'] == 'MDS':
            if 'system_image' not in comp or 'kickstart_image' not in comp:
                res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_JSON"))

                return res
        elif 'device_type' in comp and comp['device_type'] == 'UCSM':
            if 'domain_name' not in comp or 'infra_image' not in comp or 'blade_image' not in comp or 'server_type' not in comp:
                res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_JSON"))
                return res

    if not stacktype:
        res.setResult(res_data, PTK_INTERNALERROR, _("PDT_JSON_IMPORT_INVALID_JSON"))
        return res

    server_type = ''
    for comp in data['components']:
        if comp['device_type'] == 'UCSM' and 'server_type' in comp:
            server_type = comp['server_type']
    res_data['server_type'] = server_type
    res.setResult(res_data, PTK_OKAY, "success")
    return res


def set_network_details(device, global_var):
    for var in global_var:
        if var['name'] == "ntp":
            device['ntp_server'] = var["value"]
            continue

        if var['name'] == "gateway":
            device['gateway'] = var["value"]
            continue

        if var['name'] == "netmask":
            device['netmask'] = var["value"]
            continue


def json_config_defaults(stacktype):
    res = result()
    comp_list = []
    conf_list = []
    path = ''
    if os.path.exists(get_download_path() + 'import-' + stacktype + '.json') is True:
        path = get_download_path() + 'import-' + stacktype + '.json'
    elif os.path.exists(get_download_path() + 'import-' + stacktype[:-5] + '.json') is True:
        path = get_download_path() + 'import-' + stacktype[:-5] + '.json'
        # incase rack server and excel import
    loginfo(path)
    if os.path.exists(path) is True:
        tmp_data = skip_comments(path)
        data = undo_json_pretty(tmp_data)
        for comp in data['components']:
            component = {
                "device_type": comp['device_type'],
                "switch_name": comp['name'] if comp['device_type'] not in ['PURE', 'FlashBlade'] else '',
                "tag": comp['tag'] if comp['device_type'] not in ['PURE', 'FlashBlade'] else '',
                "switch_ip": comp['switch_ip'] if comp['device_type'] not in ['PURE', 'FlashBlade'] else '',
            }
            if comp['device_type'] == 'Nexus 5k':
                component['switch_image'] = {
                    "switch_system_image": comp['system_image'],
                    "switch_kickstart_image": comp['kickstart_image']}
            elif comp['device_type'] == 'Nexus 9k':
                component['switch_image'] = {
                    "switch_system_image": comp['system_image']}
            elif comp['device_type'] == 'MDS':
                component['switch_image'] = {
                    "switch_system_image": comp['system_image'],
                    "switch_kickstart_image": comp['kickstart_image']}
            elif comp['device_type'] == 'UCSM':
                component['virtual_ip'] = comp['virtual_ip']
                component['mode'] = comp['leadership']
                component['dns'] = comp['dns']
                component['domain_name'] = comp['domain_name']
                component['blade_image'] = comp['blade_image']
                component['infra_image'] = comp['infra_image']
                if comp['leadership'] == 'primary':
                    component['kvm_console_ip'] = comp['kvm_console_ip']
                    component['esxi_file'] = comp['esxi_file']
                    component['esxi_kickstart'] = comp['esxi_kickstart']
                # probable read from globls
                component['server_type'] = comp['server_type']
            elif comp['device_type'] == 'PURE':
                if comp.get('ZTP', False) is False:
                    continue
                component['array_name'] = comp['name']
                component['ct0_ip'] = comp['ct0_ip']
                component['ct1_ip'] = comp['ct1_ip']
                component['vir0_ip'] = comp['vir0_ip']
                component['dns'] = comp['dns']
                component['domain_name'] = comp['domain_name']
                component['relay_host'] = comp['relay_host']
                component['sender_domain'] = comp['sender_domain']
                component['alert_emails'] = comp['alert_emails']
                component['organization'] = comp['organization']
                component['full_name'] = comp['full_name']
                component['job_title'] = comp['job_title']
                component['timezone'] = comp['timezone']

            comp_list.append(component)
        for conf in data['global_config']:
            config = {
                "name": conf['name'],
                "value": conf['value']
            }
            conf_list.append(config)

        for comp in data['components']:
            if comp['device_type'] == 'UCSM' and comp['leadership'] == 'primary':
                conf_list.append({
                    "name": "remote_file",
                    "value": comp['esxi_file']
                })
                loginfo("added remote_file")
                break
    dt = {"devices": comp_list, "global_config": conf_list}
    res.setResult(dt, PTK_OKAY, "success")
    return res


def monitor_init_and_deploy(stacktype):
    init = 0
    while True:
        status, devices = get_xml_element(get_discovery_store(), 'validated')
        if status:
            if any(dev['configured'] == 'Re-validate' for dev in devices):
                loginfo(
                    "%s configuration failed. Please retry the configuration" % dev['name'])
                return False

            conf_list = [dev['name']
                         for dev in devices if dev['configured'] == 'Configured']
            if len(devices) == len(conf_list):
                loginfo("Initial configuration completed for %s" %
                        str(conf_list))
                init = 1
                break
            else:
                loginfo(
                    "Waiting for all devices to be configured, current:" + str(len(conf_list)))
                time.sleep(5)

    if init == 1:
        if get_config_mode() == "manual":
            loginfo("Ready for manual deployment...")
            return True
        obj = Orchestration()
        res = obj.batchexecute(stacktype, '')
        if res.getStatus() == PTK_OKAY:
            loginfo("Workflow execution started...")
            return True
        else:
            loginfo("Failed to deploy workflows. %s" % res.getMsg())
            return False
    else:
        loginfo("Initial configuration failed")
        return False


def restore_config(stacktype, datas):
    res = result()
    res_1 = save_config(stacktype, datas)
    if res_1.getStatus() == PTK_OKAY:
        res_2 = initialconfig()
        if res_2.getStatus() == PTK_OKAY:
            threading.Thread(target=monitor_init_and_deploy,
                             args=(stacktype,)).start()
            res.setResult([], PTK_OKAY, "success")
        else:
            res.setResult([], PTK_INTERNALERROR,
                          "Component Initialization Failed")
    else:
        res.setResult(res_1.getResult(), PTK_INTERNALERROR,
                      "Failed to save the configuration details")
    return res


def update_config_inputs(stacktype, jid):
    loginfo("Updating json config to xml")

    conf_json = get_download_path() + 'import-' + stacktype + '.json'
    if os.path.exists(conf_json) is True:
        # with open(conf_json, 'r') as fd:
        #    json_config = json.loads(fd.read())

        tmp_data = skip_comments(conf_json)
        json_config = undo_json_pretty(tmp_data)
    else:
        loginfo("JSON configuration file not found")
        return False

    with open(get_job_file(jid)) as td:
        jobdoc = xmltodict.parse(td.read())

    wf_list = []
    for gwf in json_config['workflows']:
        for wf in gwf['subworkflows']:
            wf_list.append(wf)

    for jwf in wf_list:
        if '@id' in jobdoc['workflow'] and jobdoc['workflow']['@id'] == jwf['wid']:
            taskset = [(xtask, jtask) for xtask in jobdoc['workflow']['tasks']['task']
                       for jtask in jwf['tasks'] if xtask['@texecid'] == jtask['texecid']]
            for task in taskset:
                task_dict = {}
                xml_task, json_task = task[0], task[1]
                xml_args = [xml_task['args']['arg']] if not isinstance(
                    xml_task['args']['arg'], list) else xml_task['args']['arg']
                argset = [(xarg, jarg) for xarg in xml_args for jarg in json_task['input']
                          if xarg['@name'] == jarg['name']]
                for arg in argset:
                    arg_dict = {}
                    xml_arg, json_arg = arg[0], arg[1]
                    arg_dict['execid'] = json_task['texecid']
                    arg_dict['ismapped'] = json_arg['mapval']
                    arg_dict['values'] = str(json_arg['value'])
                    task_dict[json_arg['name']] = arg_dict

                loginfo("Updating task id %s of workflow %s" %
                        (json_task['texecid'], jwf['wid']))
                job_task_input_save_api(
                    jid, json_task['texecid'], task_dict, "job")
    return True


def set_wf_flag(flag_list):
    doc = parse(get_skip_flag())
    wf_flags = doc.getElementsByTagName('wflag')
    for wf_flag in wf_flags:
        if wf_flag.getAttribute('name') in flag_list:
            wf_flag.setAttribute('flag', 'True')

    fd = open(get_skip_flag(), 'w')
    fd.write(pretty_print(doc.toprettyxml(indent="")))
    fd.close()


def locate_cell_by_value(sheet, search_str):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                if cell.value == search_str:
                    return {'row': cell.row, 'col': cell.column}
    return None


def get_details(sheet, locate_value, spacing):
    data = []
    ret = locate_cell_by_value(sheet, locate_value)
    if ret is None:
        return data
    global_row_start = ret['row'] + spacing
    cells = sheet['B' + str(global_row_start): 'C200']
    for c1, c2 in cells:
        if c1.value is None:
            break
        if c2.value is not None:
            data.append({'name': c1.value, 'value': c2.value})
        else:
            data.append({'name': c1.value, 'value': ''})
    return data


def add_componets_data(data, component, vals, addval):
    mds_fields = {
        'MDS System Image': 'system_image',
        'MDS Image version': 'image_version',
        'name': 'name',
        'MDS kickstart image': 'kickstart_image',
        'IP Address': 'switch_ip'}
    nexus_fields = {
        'Nexus 9k System Image': 'system_image',
        'Nexus 9k Image version': 'image_version',
        'name': 'name',
        'IP Address': 'switch_ip'}
    nexus_5k_fields = {
        'Nexus 5k System Image': 'system_image',
        'Nexus 5k Image version': 'image_version',
        'name': 'name',
        'Nexus 5k kickstart image': 'kickstart_image',
        'IP Address': 'switch_ip'}
    ucs_fields = {
        'UCS server type Blade/Rack': 'server_type',
        'name': 'name',
        'UCS Blade firmware image': 'blade_image',
        'leadership': 'leadership',
        'Nameserver IP': 'dns',
        'UCSM infra image': 'infra_image',
        'IP Address': 'switch_ip',
        'Virtual IP Address': 'virtual_ip'}
    if vals != []:
        data['components'][component] = {}
    else:
        return
    #loginfo("add_components_data" + component)
    for val in vals:
        key = val['name']
        if component.startswith('MDS'):
            key = mds_fields[key.strip()]
        elif component.startswith('Nexus 9k'):
            key = nexus_fields[key.strip()]
        elif component.startswith('Nexus 5k'):
            key = nexus_5k_fields[key.strip()]
        elif component.startswith('UCSM'):
            key = ucs_fields[key.strip()]
            data['components'][component]['domain_name'] = {
                'value': addval['domain_name']}  # is a work around fix it, bug
            data['components'][component]['dns'] = {
                'value': addval['dns']}  # is a work around fix it, bug
        else:
            loginfo("Error: Unknown val component:" + component + "key" + key)
        data['components'][component][key] = {'value': val['value']}
    if component.startswith('UCSM'):
        data['components'][component]['blade_image'] = {
            'value': data['global_config']['firmware']['value']}
        if data['components'][component]['leadership']['value'] == 'primary':
            data['components'][component]['esxi_file'] = {
                'value': data['global_config']['remote_file']['value']}
            data['components'][component]['esxi_kickstart'] = {'value': addval['kickstart']}
            data['components'][component]['kvm_console_ip'] = {'value': addval['kvm_console_ip']}


def excel_to_json(stacktype, path):
    data = {}
    data['global_config'] = {}
    data['components'] = {}
    data['version'] = {'value': "1.0"}  # dummy version
    data['tool_version'] = "1.0"  # dummy tool version
    data['workflows'] = []
    data['stacktype'] = {'value': stacktype}
    data['flags'] = {'value': ''}
    book = load_workbook(path)
    label = None
    for flashstack in g_flash_stack_types:
        if flashstack['value'] == stacktype:
            label = flashstack['label'] + "(" + flashstack['tag'] + ")"
            break
    if label is None:
        loginfo("Unable to find sutable label for the type specified")
        return None
    label = label.replace('//', '|')
    loginfo("Label=" + label)

    sheet = book[label]
    componenet_list = {
        'FlashArray': 'FLASH ARRAY CONFIGURATION',
        'MDS_A': 'MDS A CONFIGURATION',
        'MDS_B': 'MDS B CONFIGURATION',
        'Nexus 9k_A': 'NEXUS(9K) A CONFIGURATION',
        'Nexus 9k_B': 'NEXUS(9K) B CONFIGURATION',
        'UCSM_A': 'UCSM A CONFIGURATION',
        'UCSM_B': 'UCSM B CONFIGURATION',
        'Nexus 5k_A': 'NEXUS(5K) A CONFIGURATION',
        'Nexus 5k_B': 'NEXUS(5K) B CONFIGURATION'}

    global_vals = get_details(sheet, 'GLOBAL CONFIGURATIONS', 4)
    for gv in global_vals:
        if gv['name'][-1] == '*':
            gv['name'] = gv['name'][:-2]
    # print global_vals
    global_input_doc = None
    with open(get_global_wf_config_file()) as td:
        doc = xmltodict.parse(td.read())
        for htype in doc['globals']['htype']:
            if htype['@stacktype'] == stacktype:
                global_input_doc = htype['input']
                break
    addval = {}
    os_install = True
    for globalval in global_vals:
        name = None
        for input1 in global_input_doc:
            if globalval['name'] == 'Server Image':
                name = 'firmware'
                break
            if globalval['name'] == input1['@label']:
                name = input1['@name']
                break
        if name is not None:
            data['global_config'][name] = {'value': globalval['value']}
            if name == 'kvm_console_ip':
                addval['kvm_console_ip'] = globalval['value']
        else:
            if globalval['name'] == 'ESXi Kickstart':
                # an exception case
                addval['kickstart'] = globalval['value']
            elif globalval['name'] == 'OS Install':
                if globalval['value'] != 'Yes':
                    os_install = False
            elif globalval['name'] == 'Domain Name':
                addval['domain_name'] = globalval['value']
            elif globalval['name'] == 'Nameserver IP':
                addval['dns'] = globalval['value']
            else:
                loginfo("No value found for " + globalval['name'])

    if not os_install:
        data['global_config']['remote_file']['value'] = ''
        addval['kickstart'] = ''

    for key, value in componenet_list.items():
        vals = get_details(sheet, value, 4)
        for val in vals:
            if val['name'].strip()[-1] == '*':
                val['name'] = val['name'].strip()[:-2]
        add_componets_data(data, key, vals, addval)

    # correct errors in excel
    ucs_fields = ['infra_image', 'server_type', 'dns']
    for field in ucs_fields:
        data['components']['UCSM_B'][field]['value'] = data['components']['UCSM_A'][field]['value']

    if data['components']['UCSM_A']['leadership']['value'] == 'primary':
        data['components']['UCSM_B']['leadership']['value'] = 'subordinate'
    else:
        data['components']['UCSM_B']['leadership']['value'] = 'primary'

    if 'MDS_A' in data['components'].keys():
        mds_fields = ['system_image', 'kickstart_image']
        for field in mds_fields:
            data['components']['MDS_B'][field]['value'] = data['components']['MDS_A'][field]['value']

    if 'Nexus 9k_A' in data['components'].keys():
        nexus_fields = ['system_image']
        for field in nexus_fields:
            data['components']['Nexus 9k_B'][field]['value'] = data['components']['Nexus 9k_A'][field]['value']

    if 'Nexus 5k_A' in data['components'].keys():
        nexus_fields = ['system_image', 'kickstart_image']
        for field in nexus_fields:
            data['components']['Nexus 5k_B'][field]['value'] = data['components']['Nexus 5k_A'][field]['value']
    if data['global_config']['upgrade']['value'] == 'No':
        # no ucs upgrade
        data['global_config']['firmware']['value'] = ''
        data['components']['UCSM_A']['infra_image']['value'] = ''
        data['components']['UCSM_B']['infra_image']['value'] = ''
        data['components']['UCSM_B']['blade_image']['value'] = ''
        data['components']['UCSM_A']['blade_image']['value'] = ''
    return data


# Test Code
'''for st in g_flash_stack_types:
        if 'hidden' in st and st['hidden'] == True:
	    continue
	print st['value']
	print st['label']
	if st['value'] == 'fa-mds-n5k-fi-fc':
		continue
	excel_to_json(st['value'], '/tmp/SmartConfigTemplate.xlsx')
	#print json.dumps(excel_to_json(st['value'], '/tmp/SmartConfigTemplate.xlsx'))'''
#export_configuration('fa-n9k-fi-mds-fc', None)
# print data['system_info']
